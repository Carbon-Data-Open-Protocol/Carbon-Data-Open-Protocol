from pathlib import Path
from functools import lru_cache
import csv
import json
import re
from openpyxl import load_workbook


INPUT_FILE = "docs/CDOP_SCHEMA.xlsx"
ENUM_VALUES_FILE = "docs/CDOP Enumerated Values Lists.xlsx"
OUTPUT_DIR = "json_schema"
JSON_SCHEMA_VERSION = "https://json-schema.org/draft/2020-12/schema"

UNMATCHED_ENUM_FIELDS = []


TYPE_MAP = {
    "string": "string",
    "str": "string",
    "text": "string",
    "enum": "string",
    "integer": "integer",
    "int": "integer",
    "number": "number",
    "float": "number",
    "decimal": "number",
    "boolean": "boolean",
    "bool": "boolean",
    "date": "string",
    "datetime": "string",
    "object": "object",
    "array": "array",
}


FORMAT_MAP = {
    "date": "date",
    "datetime": "date-time",
}


def normalize_header(value):
    if value is None:
        return ""
    value = str(value).strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def safe_filename(value):
    value = str(value).strip()
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value)
    return value.strip("_") or "schema"


def map_type(datatype):
    if datatype is None:
        return "string"
    return TYPE_MAP.get(str(datatype).strip().lower(), "string")


def map_format(datatype):
    if datatype is None:
        return None
    return FORMAT_MAP.get(str(datatype).strip().lower())


def is_required(value):
    if value is None:
        return False
    return str(value).strip().lower() in {"required", "yes", "y", "true", "1"}


def load_enum_value_lookup(path=ENUM_VALUES_FILE):
    input_path = Path(path)

    if not input_path.exists():
        print(f"Warning: enum values file not found at {input_path}; enum fields will have no enum list")
        return {}

    wb = load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[2]
    data_rows = rows[4:]

    lookup = {}

    for col_index, header in enumerate(header_row):
        if not clean_value(header):
            continue

        values = []
        seen = set()
        for row in data_rows:
            value = clean_value(row[col_index]) if col_index < len(row) else None
            if value is not None and value not in seen:
                seen.add(value)
                values.append(value)

        if not values:
            continue

        for part in str(header).split("&"):
            key = normalize_header(part)
            if key:
                lookup[key] = values

    return lookup


@lru_cache(maxsize=None)
def get_enum_lookup():
    return load_enum_value_lookup()


def record_unmatched_enum_field(record):
    UNMATCHED_ENUM_FIELDS.append(
        (record.get("associated_entity"), record.get("field_name"))
    )


def parse_cardinality(cardinality):
    if cardinality is None:
        return {"is_array": False, "min_items": None, "max_items": None}

    value = str(cardinality).strip()

    if ".." not in value:
        return {"is_array": False, "min_items": None, "max_items": None}

    lower, upper = [part.strip() for part in value.split("..", 1)]

    min_items = int(lower) if lower.isdigit() else None
    max_items = None if upper == "*" else int(upper) if upper.isdigit() else None
    is_array = upper == "*" or (upper.isdigit() and int(upper) > 1)

    return {
        "is_array": is_array,
        "min_items": min_items,
        "max_items": max_items,
    }


def add_custom_annotations(field_schema, record):
    annotation_fields = {
        "field_id": "x-cdop-field-id",
        "pre_issuance_inclusion": "x-cdop-pre-issuance-inclusion",
        "cardinality": "x-cdop-cardinality",
        "data_source": "x-cdop-data-source",
        "mutability": "x-cdop-mutability",
        "public_private": "x-cdop-public-private",
        "schema": "x-cdop-schema",
        "parent_field": "x-cdop-parent-field",
        "item_type": "x-cdop-item-type",
        "field_path": "x-cdop-field-path",
    }

    for source_key, target_key in annotation_fields.items():
        value = record.get(source_key)
        if value is not None:
            field_schema[target_key] = value


def build_base_field_schema(record):
    datatype = record.get("datatype")
    datatype_key = str(datatype).strip().lower() if datatype else None

    field_schema = {"type": map_type(datatype)}

    json_format = map_format(datatype)
    if json_format:
        field_schema["format"] = json_format

    if record.get("definition"):
        field_schema["description"] = record["definition"]

    if datatype_key == "enum":
        field_schema["type"] = "string"
        field_name_key = normalize_header(record.get("field_name"))
        enum_values = get_enum_lookup().get(field_name_key)
        if enum_values:
            field_schema["enum"] = enum_values
        else:
            record_unmatched_enum_field(record)

    add_custom_annotations(field_schema, record)

    return field_schema


def wrap_for_cardinality(field_schema, record):
    cardinality_info = parse_cardinality(record.get("cardinality"))

    if not cardinality_info["is_array"]:
        return field_schema

    array_schema = {
        "type": "array",
        "items": field_schema,
    }

    if cardinality_info["min_items"] is not None:
        array_schema["minItems"] = cardinality_info["min_items"]

    if cardinality_info["max_items"] is not None:
        array_schema["maxItems"] = cardinality_info["max_items"]

    return array_schema


def build_field_schema(record):
    base_schema = build_base_field_schema(record)

    datatype = str(record.get("datatype")).strip().lower() if record.get("datatype") else None
    item_type = str(record.get("item_type")).strip().lower() if record.get("item_type") else None

    if datatype == "array" and item_type == "object":
        base_schema = {
            "type": "object",
            "additionalProperties": True,
            "properties": {},
            "required": [],
        }

    return wrap_for_cardinality(base_schema, record)


def ensure_object_property(parent_schema, property_name):
    props = parent_schema.setdefault("properties", {})

    if property_name not in props:
        props[property_name] = {
            "type": "object",
            "additionalProperties": True,
            "properties": {},
            "required": [],
        }

    return props[property_name]


def add_required(parent_schema, property_name):
    required = parent_schema.setdefault("required", [])
    if property_name not in required:
        required.append(property_name)
        required.sort()


def get_array_items_schema(array_schema):
    if array_schema.get("type") != "array":
        return array_schema

    items = array_schema.setdefault("items", {
        "type": "object",
        "additionalProperties": True,
        "properties": {},
        "required": [],
    })

    return items


def add_field_path_property(root_schema, field_path, record):
    """
    Supports paths like:
      project.country_code
      project.cobenefit_target[].cobenefit_type
      project.cobenefit_target[].cobenefit_value

    The first segment is treated as the top-level associated entity.
    """
    parts = [part.strip() for part in str(field_path).split(".") if part.strip()]
    if len(parts) < 2:
        raise ValueError(
            f"Invalid field_path '{field_path}' for {record.get('associated_entity')}.{record.get('field_name')} "
            f"(expected at least one '.' separator)"
        )

    current = root_schema

    for i, raw_part in enumerate(parts):
        is_last = i == len(parts) - 1
        is_array = raw_part.endswith("[]")
        part = raw_part.replace("[]", "")

        if is_last:
            current.setdefault("properties", {})[part] = build_field_schema(record)
            if is_required(record.get("required_optional")):
                add_required(current, part)
            return

        if is_array:
            current.setdefault("properties", {})
            if part not in current["properties"]:
                current["properties"][part] = {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": True,
                        "properties": {},
                        "required": [],
                    },
                }

            current = current["properties"][part]["items"]

        else:
            current = ensure_object_property(current, part)


def add_parent_field_property(root_schema, record):
    entity = normalize_header(record.get("associated_entity"))
    field_name = normalize_header(record.get("field_name"))
    parent_field = record.get("parent_field")

    entity_schema = ensure_object_property(root_schema, entity)

    if not parent_field:
        entity_schema["properties"][field_name] = build_field_schema(record)

        if is_required(record.get("required_optional")):
            add_required(entity_schema, field_name)

        return

    parent_field = normalize_header(parent_field)
    parent_schema = entity_schema["properties"].get(parent_field)

    if parent_schema is None:
        parent_schema = {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": True,
                "properties": {},
                "required": [],
            },
        }
        entity_schema["properties"][parent_field] = parent_schema

    item_schema = get_array_items_schema(parent_schema)

    item_schema.setdefault("properties", {})[field_name] = build_field_schema(record)

    if is_required(record.get("required_optional")):
        add_required(item_schema, field_name)


def schema_has_required(schema):
    """Recursively check whether a schema (or any nested property/array item) has a
    non-empty 'required' list, so top-level entities can be marked required even when
    the required fields live inside nested arrays/objects rather than directly on the
    entity itself."""
    if not isinstance(schema, dict):
        return False

    if schema.get("required"):
        return True

    for prop_schema in schema.get("properties", {}).values():
        if schema_has_required(prop_schema):
            return True

    items = schema.get("items")
    if items is not None and schema_has_required(items):
        return True

    return False


def worksheet_to_json_schema(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    headers = [normalize_header(header) for header in rows[0]]

    required_columns = {"associated_entity", "field_name", "datatype"}
    missing = required_columns - set(headers)

    if missing:
        raise ValueError(f"Sheet '{ws.title}' is missing required columns: {sorted(missing)}")

    schema = {
        "$schema": JSON_SCHEMA_VERSION,
        "$id": f"{safe_filename(ws.title)}.schema.json",
        "title": ws.title,
        "type": "object",
        "additionalProperties": True,
        "properties": {},
        "required": [],
    }

    records = []
    for row in rows[1:]:
        record = {
            headers[i]: clean_value(row[i])
            for i in range(min(len(headers), len(row)))
            if headers[i]
        }

        if record.get("associated_entity") and record.get("field_name"):
            records.append(record)

    has_field_path = any(record.get("field_path") for record in records)
    has_parent_logic = any(record.get("parent_field") or record.get("item_type") for record in records)

    if has_field_path:
        for record in records:
            field_path = record.get("field_path")
            if field_path:
                add_field_path_property(schema, field_path, record)
            else:
                add_parent_field_property(schema, record)

    elif has_parent_logic:
        parent_rows = [r for r in records if not r.get("parent_field")]
        child_rows = [r for r in records if r.get("parent_field")]

        for record in parent_rows:
            add_parent_field_property(schema, record)

        for record in child_rows:
            add_parent_field_property(schema, record)

    else:
        for record in records:
            add_parent_field_property(schema, record)

    schema["required"] = sorted(
        entity
        for entity, entity_schema in schema.get("properties", {}).items()
        if schema_has_required(entity_schema)
    )
    if not schema["required"]:
        schema.pop("required", None)

    return schema


def convert_workbook(input_file=INPUT_FILE, output_dir=OUTPUT_DIR):
    input_path = Path(input_file)
    output_path = Path(output_dir)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    output_path.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)

    for ws in wb.worksheets:
        schema = worksheet_to_json_schema(ws)
        if schema is None:
            continue

        output_file = output_path / f"{safe_filename(ws.title)}.schema.json"

        with output_file.open("w", encoding="utf-8") as f:
            json.dump(schema, f, indent=2, ensure_ascii=False)

        print(f"Wrote {output_file}")

    if UNMATCHED_ENUM_FIELDS:
        seen = set()
        deduped = []
        for entry in UNMATCHED_ENUM_FIELDS:
            if entry not in seen:
                seen.add(entry)
                deduped.append(entry)

        report_file = output_path / "unmatched_enum_fields.csv"
        with report_file.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["associated_entity", "field_name"])
            writer.writerows(deduped)

        print(f"Wrote {len(deduped)} unmatched enum fields to {report_file}")


if __name__ == "__main__":
    convert_workbook()